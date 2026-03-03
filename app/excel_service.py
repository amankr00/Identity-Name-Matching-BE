import base64
import os
import re
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Optional
from urllib.parse import parse_qs, urljoin, urlparse

import requests
from openpyxl import load_workbook
from rapidfuzz import fuzz

from .ocr_service import NameOCRExtractor, OCRExtraction

def _load_threshold() -> float:
    raw = os.getenv("MATCH_THRESHOLD", "90")
    try:
        value = float(raw)
    except ValueError:
        return 90.0
    return max(0.0, min(100.0, value))


MATCH_THRESHOLD = _load_threshold()


def _load_debug_columns_enabled() -> bool:
    raw = os.getenv("OUTPUT_DEBUG_COLUMNS", "1").strip().lower()
    return raw not in {"0", "false", "no", "off"}


DEBUG_COLUMNS_ENABLED = _load_debug_columns_enabled()


def normalize_name(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value))
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    uppercase_text = ascii_text.upper()
    letters_only = re.sub(r"[^A-Z ]", " ", uppercase_text)
    return re.sub(r"\s+", " ", letters_only).strip()


def compute_similarity(name_a: str, name_b: str) -> float:
    left = normalize_name(name_a)
    right = normalize_name(name_b)

    if not left or not right:
        return 0.0

    if left == right:
        return 100.0

    return float(
        max(
            fuzz.ratio(left, right),
            fuzz.partial_ratio(left, right),
            fuzz.token_sort_ratio(left, right),
        )
    )


def compute_best_similarity_in_text(reference_name: str, ocr_text: str) -> float:
    reference = normalize_name(reference_name)
    full_text = normalize_name(ocr_text)
    if not reference or not full_text:
        return 0.0

    text_tokens = full_text.split()[:80]
    reference_tokens = reference.split()
    if len(reference_tokens) >= 2 and all(token in text_tokens for token in reference_tokens):
        return 100.0

    best = compute_similarity(reference, full_text)
    if not reference_tokens or not text_tokens:
        return best

    token_count = len(reference_tokens)
    min_size = max(2, token_count - 1)
    max_size = min(len(text_tokens), token_count + 1)

    for size in range(min_size, max_size + 1):
        for start in range(0, len(text_tokens) - size + 1):
            candidate = " ".join(text_tokens[start : start + size])
            score = compute_similarity(reference, candidate)
            if score > best:
                best = score

    return best


def _embedded_images_by_row(worksheet) -> dict[int, bytes]:
    row_to_image: dict[int, bytes] = {}
    for image in getattr(worksheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        if anchor is None or not hasattr(anchor, "_from"):
            continue

        try:
            row = anchor._from.row + 1
            col = anchor._from.col + 1
        except Exception:
            continue

        if col != 2:
            continue

        try:
            row_to_image[row] = image._data()
        except Exception:
            continue

    return row_to_image


def _decode_data_url(value: str) -> Optional[bytes]:
    if not value.lower().startswith("data:image"):
        return None
    if "," not in value:
        return None

    payload = value.split(",", 1)[1]
    return base64.b64decode(payload)


def _extract_hyperlink_formula_url(value: str) -> Optional[str]:
    match = re.match(r'^\s*=HYPERLINK\(\s*"([^"]+)"\s*,', value, flags=re.IGNORECASE)
    if not match:
        return None
    return match.group(1).strip()


def _resolve_image_source(cell) -> Optional[str]:
    if getattr(cell, "hyperlink", None) is not None:
        target = getattr(cell.hyperlink, "target", None)
        if isinstance(target, str) and target.strip():
            return target.strip()

    raw_value = cell.value
    if not isinstance(raw_value, str):
        return None

    text = raw_value.strip()
    if not text:
        return None

    formula_url = _extract_hyperlink_formula_url(text)
    if formula_url:
        return formula_url

    return text


def _normalize_google_drive_url(url: str) -> str:
    parsed = urlparse(url)
    host = parsed.netloc.lower()
    if "drive.google.com" not in host:
        return url

    path_match = re.search(r"/file/d/([^/]+)", parsed.path)
    if path_match:
        file_id = path_match.group(1)
        return f"https://drive.google.com/uc?export=download&id={file_id}"

    query_id = parse_qs(parsed.query).get("id", [])
    if query_id:
        file_id = query_id[0]
        return f"https://drive.google.com/uc?export=download&id={file_id}"

    return url


def _extract_image_url_from_html(base_url: str, html: str) -> Optional[str]:
    patterns = [
        r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']',
        r'<img[^>]+src=["\']([^"\']+)["\']',
    ]

    for pattern in patterns:
        match = re.search(pattern, html, flags=re.IGNORECASE)
        if not match:
            continue
        candidate = match.group(1).strip()
        if candidate:
            return urljoin(base_url, candidate)

    return None


def _download_url_image(url: str) -> tuple[Optional[bytes], str]:
    normalized_url = _normalize_google_drive_url(url)
    try:
        response = requests.get(
            normalized_url,
            timeout=20,
            allow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 (ID-Name-Verification/1.0)"},
        )
        response.raise_for_status()
    except requests.RequestException as exc:
        return None, f"url_fetch_error:{type(exc).__name__}"

    content = response.content
    if not content:
        return None, "url_fetch_error:empty_response"

    content_type = response.headers.get("content-type", "").lower()
    if "image/" in content_type:
        return content, "url_image"

    try:
        html = content.decode("utf-8", errors="ignore")
    except Exception:
        return content, f"url_non_image_content_type:{content_type or 'unknown'}"

    discovered = _extract_image_url_from_html(response.url, html)
    if not discovered:
        return content, f"url_non_image_content_type:{content_type or 'unknown'}"

    try:
        follow_up = requests.get(
            discovered,
            timeout=20,
            allow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 (ID-Name-Verification/1.0)"},
        )
        follow_up.raise_for_status()
        follow_up_content = follow_up.content
        if follow_up_content:
            return follow_up_content, "url_image_from_html"
    except requests.RequestException as exc:
        return None, f"url_nested_fetch_error:{type(exc).__name__}"

    return None, "url_nested_fetch_error:empty_response"


def _image_from_cell_value(value: Optional[str]) -> tuple[Optional[bytes], str]:
    if not isinstance(value, str):
        return None, "missing_source"

    raw_value = value.strip()
    if not raw_value:
        return None, "missing_source"

    data_url_image = _decode_data_url(raw_value)
    if data_url_image:
        return data_url_image, "data_url"

    if raw_value.lower().startswith(("http://", "https://")):
        return _download_url_image(raw_value)

    candidate_path = Path(raw_value).expanduser()
    if candidate_path.exists() and candidate_path.is_file():
        return candidate_path.read_bytes(), "local_path"

    return None, "unsupported_source"


def _build_debug_reason(
    source_state: str,
    ocr_result: OCRExtraction,
    extracted_name: str,
    final_similarity: float,
) -> str:
    if source_state == "embedded_image":
        source_state = "embedded_image"

    if source_state.startswith(("url_fetch_error", "url_nested_fetch_error")):
        return source_state

    if source_state in {"missing_source", "unsupported_source"}:
        return source_state

    if ocr_result.engine_name == "none":
        return "ocr_engine_missing"

    if not ocr_result.raw_text.strip():
        return "ocr_no_text"

    if not extracted_name:
        return f"name_not_detected:score={final_similarity:.1f}"

    return f"low_similarity:score={final_similarity:.1f}"


def process_workbook(file_bytes: bytes) -> bytes:
    workbook = load_workbook(filename=BytesIO(file_bytes))
    worksheet = workbook.active
    ocr_extractor = NameOCRExtractor()

    image_map = _embedded_images_by_row(worksheet)
    worksheet.cell(row=1, column=3, value="Match Status")
    if DEBUG_COLUMNS_ENABLED:
        worksheet.cell(row=1, column=4, value="Extracted Name")
        worksheet.cell(row=1, column=5, value="Similarity Score")
        worksheet.cell(row=1, column=6, value="Debug")
        worksheet.cell(row=1, column=7, value="OCR Preview")

    for row in range(2, worksheet.max_row + 1):
        input_name = worksheet.cell(row=row, column=1).value
        image_cell = worksheet.cell(row=row, column=2)
        image_cell_value = _resolve_image_source(image_cell)

        reference_name = str(input_name).strip() if input_name is not None else ""
        image_bytes = image_map.get(row)
        source_state = "embedded_image" if image_bytes is not None else "missing_source"
        if image_bytes is None:
            try:
                image_bytes, source_state = _image_from_cell_value(image_cell_value)
            except Exception:
                image_bytes = None
                source_state = "image_source_error"

        ocr_result = (
            ocr_extractor.extract_details(image_bytes)
            if image_bytes
            else OCRExtraction(engine_name=ocr_extractor.engine_name, raw_text="", extracted_name="")
        )
        extracted_name = ocr_result.extracted_name

        similarity_from_name = compute_similarity(reference_name, extracted_name)
        similarity_from_text = compute_best_similarity_in_text(reference_name, ocr_result.raw_text)
        similarity = max(similarity_from_name, similarity_from_text)

        status = "Matched" if similarity >= MATCH_THRESHOLD else "Not Matched"
        debug_reason = (
            "matched"
            if status == "Matched"
            else _build_debug_reason(source_state, ocr_result, extracted_name, similarity)
        )
        ocr_preview = normalize_name(ocr_result.raw_text)[:140]

        worksheet.cell(row=row, column=3, value=status)
        if DEBUG_COLUMNS_ENABLED:
            worksheet.cell(row=row, column=4, value=extracted_name)
            worksheet.cell(row=row, column=5, value=round(similarity, 2))
            worksheet.cell(row=row, column=6, value=debug_reason)
            worksheet.cell(row=row, column=7, value=ocr_preview)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()
